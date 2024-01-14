from cProfile import label
from cgitb import text
from ctypes import sizeof
from logging import root
from struct import pack
from tkinter import *
from email.mime import image
from encodings import utf_8
from tkinter import dialog
from tracemalloc import start
from typing import List
from unittest import expectedFailure
from openpyxl import workbook, load_workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image 
import os
import time
import shutil
import webbrowser
import datetime

local_icone_janela = 'ms-excel-icon.ico'
texto_retorno = ''
font_retorno = 'Arial Black'

txt_config = open('configs.txt', 'r')
cor_layout_tela = txt_config.readline()


def Finalizando_Criacao_Planilha():

        #arq_plani = open('PLANILHAS.txt', 'r', encoding='utf-8')
        #arq_plani = [s.rstrip() for s in arq_plani]

    try:    

        arq_plani = nova_planilhas_nome.get()
        arq_plani = [x for x in arq_plani.split(',')]

        mes_atual_planilha = str(info_mesatual.get())
        
        for i in range(len(arq_plani)):

            nome_nova_planilha = arq_plani[i]
            wb = load_workbook('modelo.xlsx')

            img = Image('tvco_logo.png')
            img.height = 85 
            img.width = 95
            img2 = Image('sbt_logo.png')
            img2.height = 85 
            img2.width = 90
            ws = wb.active
            ws.add_image(img,'A3')
            ws.add_image(img2,'B3')

            ws['K5'] = nome_nova_planilha
            ws['K7'] = mes_atual_planilha
            wb.save(arq_plani[i]+'.xlsx')

            arq_criados_retorno['text'] = 'Criado com sucesso!'
    except:
        arq_criados_retorno['text'] = 'Por favor, Preencha o campo corretamente.'

def limpa_dados_novaplanilha():
    arq_criados_retorno['text'] = 'Ex: Loja_1,Loja2,Posto10'

def limpadados_horario():
    ad_horarios_retorno['text'] = texto_retorno 

def limpar_no_deletar():
    Exc_planilhas_retorno['text'] = texto_retorno

def limpa_bakcup_retorno():
    retorno_backup['text'] = texto_retorno

def Nova_Planilha():

        janela_Criando_Planilha = Tk()
        janela_Criando_Planilha.title('Go - NOVA PLANILHA')
        janela_Criando_Planilha.geometry('300x300')
        janela_Criando_Planilha.resizable(0,0)
        janela_Criando_Planilha.configure(background=cor_layout_tela)

        janela_Criando_Planilha.iconbitmap(local_icone_janela)

        Label(janela_Criando_Planilha, text='MÊS E ANO', bg=cor_layout_tela, fg='White', 
        font='Arial').place(x=75, y=10, width=150,height=30)

        global info_mesatual 
        info_mesatual = Entry(janela_Criando_Planilha)
        mes_automatico_atual = datetime.datetime.now()

        if mes_automatico_atual.month == 1:
            nome_mes_planilha = 'JANEIRO'
        elif mes_automatico_atual.month == 2:
            nome_mes_planilha = 'FEVEREIRO'
        elif mes_automatico_atual.month == 3:
            nome_mes_planilha = 'MARÇO'
        elif mes_automatico_atual.month == 4:
            nome_mes_planilha = 'ABRIL'
        elif mes_automatico_atual.month == 5:
            nome_mes_planilha = 'MAIO'
        elif mes_automatico_atual.month == 6:
            nome_mes_planilha = 'JUNHO'
        elif mes_automatico_atual.month == 7:
            nome_mes_planilha = 'JULHO'
        elif mes_automatico_atual.month == 8:
            nome_mes_planilha = 'AGOSTO'
        elif mes_automatico_atual.month == 9:
            nome_mes_planilha = 'SETEMBRO'
        elif mes_automatico_atual.month == 10:
            nome_mes_planilha = 'OUTUBRO'
        elif mes_automatico_atual.month == 11:
            nome_mes_planilha = 'NOVEMBRO'
        elif mes_automatico_atual.month == 12:
            nome_mes_planilha = 'DEZEMBRO'
        else:
            nome_mes_planilha = 'Mes_inexistente'

        real = nome_mes_planilha,'-',mes_automatico_atual.year
        info_mesatual.insert(10,real)
        info_mesatual.place(x=75, y=45, width=150,height=30)

        botao_criar = Button(janela_Criando_Planilha, text='CRIAR', bg='White', 
        font=('Arial',10), command=Finalizando_Criacao_Planilha).place(x=10, y=90, width=130,height=30)

        botao_limpar = Button(janela_Criando_Planilha, text='LIMPAR', bg='White',
        font=('Arial',10), command=limpa_dados_novaplanilha).place(x=160, y=90, width=130,height=30)

        global nova_planilhas_nome
        nova_planilhas_nome = Entry(janela_Criando_Planilha, bd=3)
        nova_planilhas_nome.insert(10, 'nome_da_planilha_aqui,')
        nova_planilhas_nome.place(x=10, y=140, width=280,height=60)

        global arq_criados_retorno 
        arq_criados_retorno = Label(janela_Criando_Planilha, text='Ex: Loja_1,Loja2,Posto10', 
        fg='White', bg=cor_layout_tela,font=(font_retorno, 10))
        arq_criados_retorno.place(x=10, y=220, width=280,height=45)

        Label(janela_Criando_Planilha, text='Desenvolvido por Rodrigo', bg=cor_layout_tela, fg='Black', 
        font=('Arial', 7)).place(x=30, y=270, width=240,height=30)

        janela_Criando_Planilha.mainloop()

def Adc_finalizando():

        #arquivo = open('BLOCO.txt', 'r', encoding='utf-8')
        #arquivo = [s.rstrip() for s in arquivo]

    try:
        arquivo = planilhas_entrada.get()
        arquivo = [x for x in arquivo.split(',')]

        dia_mes = int(dia_atual.get())
        dia_mes = dia_mes + 10

        for i in range(len(arquivo)):
            wb = load_workbook(arquivo[i]+'.xlsx')
            ws = wb.active

            verificador_de_dados = 0

            for x in range(15):
                verificador_de_dados = verificador_de_dados + 1

                if ws.cell(dia_mes ,verificador_de_dados).value == None:
                    ws.cell(dia_mes ,verificador_de_dados).value = horario_atual.get()
                    wb.save(arquivo[i]+'.xlsx')
                    ad_horarios_retorno['text'] = 'Horário adicionado na planilha!'
                    break
    except:
        ad_horarios_retorno['text'] = 'Algum dado está incorreto.'

def Adicionando_horario():  
        janela_Adicionando_Horario = Tk()
        janela_Adicionando_Horario.title('Go - HORÁRIO')
        janela_Adicionando_Horario.geometry('300x300')
        janela_Adicionando_Horario.resizable(0,0)
        janela_Adicionando_Horario.configure(background=cor_layout_tela)  

        janela_Adicionando_Horario.iconbitmap(local_icone_janela)

        Label(janela_Adicionando_Horario, text='DIA:', bg=cor_layout_tela, fg='White', 
        font=('Arial',10)).place(x=10, y=20, width=25,height=30)

        global dia_atual
        dia_atual = Entry(janela_Adicionando_Horario)
        day_hoje = datetime.datetime.now()
        dia_atual.insert(10,day_hoje.day)
        dia_atual.place(x=40, y=20, width=80,height=30)

        Label(janela_Adicionando_Horario, text='HORÁRIO:', bg=cor_layout_tela, fg='White', 
        font=('Arial',10)).place(x=140, y=20, width=60,height=30)

        global horario_atual
        horario_atual = Entry(janela_Adicionando_Horario)
        horario_atual.place(x=210, y=20, width=80,height=30)

        botao_add_horario = Button(janela_Adicionando_Horario, text='ADICIONAR', bg='White', 
        font=('Arial',10), command=Adc_finalizando).place(x=35, y=80, width=110,height=30)

        botao_limpadados_horario = Button(janela_Adicionando_Horario, text='LIMPAR', bg='White', 
        font=('Arial',10), command=limpadados_horario).place(x=160, y=80, width=110,height=30)

        global planilhas_entrada
        planilhas_entrada = Entry(janela_Adicionando_Horario, bd=3)
        planilhas_entrada.insert(10, 'nome_da_planilha_aqui')
        planilhas_entrada.place(x=10, y=140, width=280,height=60)

        global ad_horarios_retorno 
        ad_horarios_retorno = Label(janela_Adicionando_Horario, text=texto_retorno, fg='White', bg=cor_layout_tela,
        font=(font_retorno, 10))
        ad_horarios_retorno.place(x=10, y=220, width=280,height=45)

        Label(janela_Adicionando_Horario, text='Desenvolvido por Rodrigo', bg=cor_layout_tela, fg='Black', 
        font=('Arial', 7)).place(x=30, y=270, width=240,height=30)
   
        janela_Adicionando_Horario.mainloop()

def finali_delete_planilha():
    try:
        nome_arquivo_deletar = plani_delete.get()
        if os.path.exists(nome_arquivo_deletar+'.xlsx'):
            os.remove(nome_arquivo_deletar+'.xlsx')
            Exc_planilhas_retorno['text'] = 'A planilha foi deletada!'        
    except:
            Exc_planilhas_retorno['text'] = 'Planilha não encontrada!'

def Excluindo_planilha():
    janela_Excluindo_planilha = Tk()
    janela_Excluindo_planilha.title('Go - DELETAR')
    janela_Excluindo_planilha.geometry('300x250')
    janela_Excluindo_planilha.resizable(0,0)
    janela_Excluindo_planilha.configure(background=cor_layout_tela)

    janela_Excluindo_planilha.iconbitmap(local_icone_janela)

    Label(janela_Excluindo_planilha, text='NOME DA PLANILHA:', bg=cor_layout_tela, fg='White', 
    font=('Arial',10)).place(x=10, y=20, width=140,height=30)

    global plani_delete
    plani_delete = Entry(janela_Excluindo_planilha)
    plani_delete.place(x=150, y=20, width=135,height=30)

    botao_delete = Button(janela_Excluindo_planilha, text='DELETAR', bg='White',
    font=('Arial',10), command=finali_delete_planilha).place(x=30, y=90, width=110,height=30)

    botao_limpa_deletar = Button(janela_Excluindo_planilha, text='LIMPAR', bg='White',
    font=('Arial',10), command=limpar_no_deletar).place(x=160, y=90, width=110,height=30)

    global Exc_planilhas_retorno 
    Exc_planilhas_retorno = Label(janela_Excluindo_planilha, text=texto_retorno, fg='White', bg=cor_layout_tela,
    font=(font_retorno, 10))
    Exc_planilhas_retorno.place(x=20, y=160, width=260,height=45)

    Label(janela_Excluindo_planilha, text='Desenvolvido por Rodrigo', bg=cor_layout_tela, fg='Black', 
    font=('Arial', 7)).place(x=30, y=225, width=240,height=30)

    janela_Excluindo_planilha.mainloop()

def fazendo_backup():
    try:
        end_raiz = de_pastaraiz.get()
        end_destino = para_pastadestino.get()

        shutil.copy(end_raiz, end_destino)
        retorno_backup['text'] = 'Backup feito com sucesso!'

    except:
        retorno_backup['text'] = 'Por favor, Preencha os campos corretamente.'

def salve_ende_backup():
    try:
        arquivo = open('save_backup.txt', 'w')
        arquivo.write(de_pastaraiz.get())
        arquivo.write('\n')
        arquivo.write(para_pastadestino.get())
        arquivo.close()

        retorno_backup['text'] = 'Endereço salvo!'

    except:
        retorno_backup['text'] = 'Endereço não salvo!'

def Backup_planilha():
    janela_Backup_planilha = Tk()
    janela_Backup_planilha.title('Go - BACKUP')
    janela_Backup_planilha.geometry('300x250')
    janela_Backup_planilha.resizable(0,0)
    janela_Backup_planilha.configure(background=cor_layout_tela)

    janela_Backup_planilha.iconbitmap(local_icone_janela)

    arquivo = open('save_backup.txt', 'r')
    primeira_linha = arquivo.readline()
    segunda_linha = arquivo.readline()

    Label(janela_Backup_planilha, text='DE:', bg=cor_layout_tela, fg='White', 
    font=('Arial',10)).place(x=0, y=20, width=60,height=30)
    
    global de_pastaraiz 
    de_pastaraiz = Entry(janela_Backup_planilha)
    de_pastaraiz.insert(10,primeira_linha)
    de_pastaraiz.place(x=50, y=20, width=230,height=30)

    Label(janela_Backup_planilha, text='PARA:', bg=cor_layout_tela, fg='White', 
    font=('Arial',10)).place(x=0, y=80, width=60,height=30)

    global para_pastadestino
    para_pastadestino = Entry(janela_Backup_planilha)
    para_pastadestino.insert(10,segunda_linha)
    para_pastadestino.place(x=60, y=80, width=220,height=30)

    botao_fazer_backup = Button(janela_Backup_planilha, text='BACKUP', bg='White',
    font=('Arial',10), command=fazendo_backup).place(x=30, y=140, width=110,height=30)

    botao_limpar_backup = Button(janela_Backup_planilha, text='LIMPAR', bg='White',
    font=('Arial',10), command=limpa_bakcup_retorno).place(x=160, y=140, width=110,height=30)

    botao_save_end = Button(janela_Backup_planilha, text='Manter', bg='White', font=('Arial',8), bd=0,
    command=salve_ende_backup).place(x=240, y=58, width=40,height=15)

    global retorno_backup
    retorno_backup = Label(janela_Backup_planilha, text=texto_retorno, fg='White', bg=cor_layout_tela,
    font=(font_retorno,8))
    retorno_backup.place(x=1, y=195, width=300,height=20)

def redi_botao_versao():
    new=2
    url= 'https://app.mediafire.com/sjuix1buinp23'
    webbrowser.open(url,new=new)

def created_by():
    new=2
    url= 'https://www.instagram.com/rodrigo_almeida_19'
    webbrowser.open(url,new=new)

def aplicando_configs_tela():

    txt_config2 = open('configs.txt', 'w')
    txt_config2.write(entrada_cor.get())

    if entrada_cor.get() == None:
        texto_retorno_configs['text'] = 'Preencha os campos.'       
    else:  
        texto_retorno_configs['text'] = 'Aplicado com sucesso, Reinicie o programa.'     
    
    txt_config2.close()

def restaurando_configs():

    txt_config3 = open('configs.txt', 'w')
    txt_config3.write('Arial Black')
    txt_config3.write('\n')
    txt_config3.write('#2E8B57')
    txt_config3.close()

    texto_retorno_configs['text'] = 'Restaurado com sucesso, Reinicie o programa.'

def config_tela():
    tela_config = Tk()
    tela_config.title('Configurações')
    tela_config.geometry('400x400')
    tela_config.resizable(0,0)
    tela_config.configure(background=cor_layout_tela)
    tela_config.iconbitmap(local_icone_janela)

    Label(tela_config, text='Cor principal', bg=cor_layout_tela, fg='White', 
    font=('Arial Black', 10), ).place(x=10, y=10, width=90,height=30)
    global entrada_cor
    entrada_cor = Entry(tela_config, bd=1)
    entrada_cor.insert(10,cor_layout_tela)
    entrada_cor.place(x=110, y=10, width=100,height=30)

    botao_aplicar = Button(tela_config, text='Aplicar', bg='White', fg='Black', 
    font=('Arial',10), command=aplicando_configs_tela).place(x=180, y=350, width=95,height=30)

    botao_cancelar = Button(tela_config, text='Restaurar', bg='White', fg='Black', 
    font=('Arial',10), command=restaurando_configs).place(x=280, y=350, width=95,height=30)

    global texto_retorno_configs
    texto_retorno_configs = Label(tela_config, text='', 
    bg=cor_layout_tela, fg='White', font=(font_retorno, 10))
    texto_retorno_configs.place(x=1, y=300, width=400,height=30)

try:

    janela_Inicial = Tk()
    janela_Inicial.title('exGo')
    janela_Inicial.geometry('240x280')
    janela_Inicial.resizable(0,0)
    janela_Inicial.configure(background=cor_layout_tela)
    janela_Inicial.iconbitmap(local_icone_janela)

    Label(janela_Inicial, text='exGo', bg=cor_layout_tela, fg='White', 
    font=('Arial', 30), ).place(x=0, y=15, width=240,height=30)

    Button(janela_Inicial, text='Desenvolvido por Rodrigo',bd=0, bg=cor_layout_tela, fg='Black', 
    font=('Arial', 7), command=created_by).place(x=0, y=250, width=240,height=30)

    Button(janela_Inicial, text='v2.3', bg=cor_layout_tela, fg='Black', bd=0,
    font=('Arial', 7),command=redi_botao_versao).place(x=165, y=30, width=30,height=30)

    botao_opcao1 = Button(janela_Inicial, text='NOVA PLANILHA', bg='White', fg='Black',font=('Arial',10),
    command=Nova_Planilha).place(x=45, y=70, width=150,height=30)

    botao_opcao2 = Button(janela_Inicial, text='HORÁRIO', bg='White', fg='Black', 
    font=('Arial',10), command=Adicionando_horario).place(x=45, y=110, width=150,height=30)

    botao_opcao3 = Button(janela_Inicial, text='DELETAR', bg='White', fg='Black', 
    font=('Arial',10), command=Excluindo_planilha).place(x=45, y=150, width=150,height=30)

    botao_opcao3 = Button(janela_Inicial, text='BACKUP', bg='White', fg='Black', 
    font=('Arial',10), command=Backup_planilha).place(x=45, y=190, width=150,height=30)

    bota_config = Button(janela_Inicial, text= '+', bg=cor_layout_tela, fg='Black', bd=0,
    font=('Arial', 12),command=config_tela).place(x=1, y=1, width=20,height=20)

except:

    txt_config3 = open('configs.txt', 'w')
    txt_config3.write('#2E8B57')
    txt_config3.close()

janela_Inicial.mainloop()


#MODIFICAÇÃO MAIS RECENTE (23/04/22) V2.2:
#OK ADD RECONHECIMENTO DOS NOMES POR SEPARAÇÃO ',' (['TESTE'], ['OUTRO TESTE']) NOMES DAS PLANILHAS NA LISTA.
#OK MELHORIAS NO RETORNO AO USUARIO
#OK MELHORIAS NAS BORDAS DO CAMPO DE PREENCHIMENTO
#OK MELHORIA RETIRADA DA BORDA BOTÃO SALVAR ENDEREÇO
#OK ADD BOTÃO DE REDIRECIONAMENTO PARA DESCRIÇÃO DO APLICATIVO /OU ALGO CRIATIVO.
#OK CRIAÇÃO DE VARIAVEIS GLOBAIS PARA MODIFICAÇÃO DE TEXTOS E FONTS
#OK SOLUÇÃO PARA ERROS DE DADOS NA INICILIZAÇÃO DA JANELA (RECUPERAÇÃO)
#OK CRIAÇÃO DE ABA DE CONFIGURAÇÕES
#OK CORREÇAO DE BUGS

#MODIFICAÇÃO (25/04/2022) V2.3:

#OK COLOCAR MES/ANO AUTOMATICO NO CAMPO DE CRIAÇÃO DE PLANILHA
#OK BOTAO DE REDIRECIONAMENTO PERFIL RODRIGO NOS CRÉDITOS INICIAL